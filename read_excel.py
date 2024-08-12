import openpyxl

path = r"E:\Desktop\UBA SOLUTION\postman exercise1.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active  #get_sheet_by_name("sheet1")

rows = sheet.max_row
column = sheet.max_column

print(rows)
print(column)

for r in range(1,rows+1):
    for c in range(1, column + 1):
        print(sheet.cell(row=r, column=c).value, end= "  ")

    print()
